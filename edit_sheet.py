import openpyxl as xl
import os
from openpyxl.chart import BarChart, Reference
from os.path import isfile
import pandas as pd


def create_new_xl_file():
    filename = input('Enter xl file name : ')
    xl_extention = '.xlsx'
    new_xl_extension = filename[-5:]

    if new_xl_extension != xl_extention:
        filename = filename + xl_extention

    sheetname = input('Enter Sheet name : ')
    wb = xl.Workbook()
    ws = wb.active
    ws.title = sheetname
    wb.save(filename)


def create_new_sheet():
    filename = 'abc.xlsx'
    if isfile(filename):
        sheet_name = input('Enter new sheet name : ')
        wb = xl.load_workbook('abc.xlsx')
        wb.create_sheet(sheet_name)
        wb.save(filename)
    else:
        print('File does not exist')


def view_xl_file():
    print("""
1. View File
2. Open File
3. Main Menu
    """)
    option = input('Chose your option : ')
    if option == '1':
        excel_data_df = pd.read_excel('xlsheet.xlsx', sheet_name = 'Sheet1')
        print(excel_data_df)
    elif option == '2':
        os.startfile('xlsheet.xlsx')


def update_price():
    filename = 'xlsheet.xlsx'
    change_rate = input('Enter change rate: ')
    change_rate = float(change_rate)

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    #cell = sheet['A1']
    cell = sheet.cell(1,1)
    print(cell.value)
    print(sheet.max_row)

    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrceted_price = cell.value * change_rate
        corrcected_price_cell = sheet.cell(row, 4)
        corrcected_price_cell.value = corrceted_price

    values = Reference(sheet,
                       min_row = 2,
                       max_row = sheet.max_row,
                       min_col = 3,
                       max_col = 3)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'F1')

    values1 = Reference(sheet,
                       min_row = 2,
                       max_row = sheet.max_row,
                       min_col = 4,
                       max_col = 4)
    chart1 = BarChart()
    chart1.add_data(values1)
    sheet.add_chart(chart1, 'F16')
    wb.save(filename)

def edit_data():
    file_name = 'xlsheet.xlsx'
    sheet_name = 'Sheet1'
    wb = xl.load_workbook(file_name)
    sheet = wb[sheet_name]
    print("""
1. Edit Cell
2. Edit row 
3. Edit column
    """)
    option = input('Chose your option : ')
    if option == '1':
        get_cell = input('Enter cell : ')
        get_value = int(input('Enter value to update :'))
        cell = sheet[get_cell]
        cell.value = get_value
        wb.save(file_name)
    elif option == '2':
        print("print")
    elif option == '3':
        print("print")
